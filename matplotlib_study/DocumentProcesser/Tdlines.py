# 加载曲线数据
# 将曲线数据写入txt?
# 将曲线数据写入xlsx?
import json
import openpyxl
import time
from models.td import Td
import matplotlib.pyplot as plt  # 导入绘图包
import matplotlib as mpl
import numpy as np
from utils import log


def open_xlsx(filename='', sheetname=''):
    # 打开xlsx文件
    # 返回一个工作表
    # filename是文件名 sheetname 是工作表名字

    filename = '三条tD曲线数据.xlsx'
    sheetname = '三条曲线数据' # 如何或区域excel中所有sheet的名字

    workbook = openpyxl.load_workbook(filename)
    sheet = workbook[sheetname]

    # 所有的工作表单名字
    # log(workbook.sheetnames)
    # 拿到某个 工作表 内容的标准做法，我们这个 excel 文件只有一个三条曲线数据工作表

    return workbook
    pass


def open_sheet(wb, sname=''):
    # 接收一个工作簿, 工作表名
    # 返回表中所有的数据
    # filename是文件名 sheetname 是工作表名字
    # sheetname = '三条曲线数据' # 如何或区域excel中所有sheet的名字

    workbook = wb
    sheetname = sname

    sheet = workbook[sheetname]

    max_column = sheet.max_column      # 最大列数
    max_row = sheet.max_row         # 最大行数

    log('max_column ({}), max_row({})'.format(max_column, max_row))


def load_data_from_xlsx():
    # 从xlsx文件中加载曲线数据
    filename = '三条tD曲线数据.xlsx'
    wb = open_xlsx(filename)
    sheetname = '三条曲线数据' # 如何或区域excel中所有sheet的名字
    sheet = open_sheet(wb, sheetname)
    pass


def test():
    load_data_from_xlsx()
    pass


if __name__ == '__main__':
    test()


